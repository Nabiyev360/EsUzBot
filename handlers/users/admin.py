from aiogram import types
from aiogram.dispatcher import FSMContext
from data.config import ADMINS
from openpyxl import load_workbook
import asyncio

from loader import dp, db
from states.adminStates import AdminState
from keyboards.inline.adminCancel import admin_cancel


@dp.message_handler(user_id=ADMINS, text=['/help_admin'])
async def send_message_users(message: types.Message):
    await message.answer(f"/count - Foydalanuvchilar soni\n\
/count_group - Bot qo'shilgan guruhlar soni\n\
/message - Foydalanuvchilarga xabar yuborish\n\
/get_words - Bazadagi so'zlar excel formatda\n\
/add_channel - Majburiy a'zolik kanal qo'shish\n\
/example - So'z qo'shish uchun misol")


# Reklama yuborish
@dp.message_handler(user_id=ADMINS, text="/message")
async def send_message_users(message: types.Message):
    await message.answer(
        text="Foydalanuvchilarga yuborish kerak bo'lgan xabarni yuboring🔼\n\nBekor qilish uchun \"Cancel\"ni bosing",
        reply_markup=admin_cancel)
    await AdminState.waiting_admin_message.set()


@dp.message_handler(user_id=ADMINS, text="/count")
async def send_count_users(message: types.Message):
    await message.answer(db.count_users())


@dp.message_handler(user_id=ADMINS, text="/count_group")
async def send_count_users(message: types.Message):
    await message.answer(db.count_group())


# send database file to admin
@dp.message_handler(user_id=ADMINS, text="/get_words")
async def send_datafile(message: types.Message):
    file = open('words_db.xlsx', 'rb')
    await message.answer_document(file)
    file.close()


@dp.message_handler(user_id=ADMINS, text="/get_users")
async def send_datafile(message: types.Message):
    file = open('data/main.db', 'rb')
    await message.answer_document(file)
    file.close()


@dp.message_handler(user_id=ADMINS, text="/example")
async def send_example(message: types.Message):
    await message.answer_photo(photo='https://telegra.ph/file/8b855b7330d14ddd64694.jpg')


@dp.message_handler(user_id=ADMINS, text="/add_channel")
async def send_message_users(message: types.Message):
    await message.answer(
        text="Majburiy a'zolikga qo'shmoqchi bo'lgan kanalingizdan biron postni menga forward qiling",
        reply_markup=admin_cancel)
    await AdminState.waiting_admin_channel.set()


@dp.message_handler(user_id=ADMINS, state=AdminState.waiting_admin_message, content_types=types.ContentType.ANY)
async def send_message_users(message: types.Message, state: FSMContext):
    await state.finish()
    await message.answer('Xabar foydalanuvchilarga yuborilmoqda...')
    n = 0
    for i in db.get_users_id():
        user_id = i[0]
        try:
            await message.send_copy(chat_id=user_id)
            n += 1
        except:
            pass
        await asyncio.sleep(0.5)
    await message.answer(f'Xabar {n} ta foydalanuvchiga muvaffaqqiyatli yuborildi!')


@dp.callback_query_handler(user_id=ADMINS, state='*', text='cancel_admin')
async def cancel_send_msg(call: types.CallbackQuery, state: FSMContext):
    await state.finish()
    await call.message.delete()
    await call.message.answer("Bekor qilindi!")


@dp.message_handler(user_id=ADMINS, state=AdminState.waiting_admin_channel, content_types=types.ContentTypes.ANY)
async def input_channel(message: types.Message, state: FSMContext):
    channel_id = message.forward_from_chat.id
    channel_name = message.forward_from_chat.title
    channel_username = message.forward_from_chat.username

    db.add_channel(channel_id, channel_name, channel_username)
    await message.answer(f"{channel_name} kanali majburiy a'zolik ro'yxatiga qo'shildi!")
    await state.finish()


# Add New Words
@dp.message_handler(user_id=ADMINS, state='*')
async def word_adder(message: types.Message, state=FSMContext):
    await state.finish()
    if message.text[0] == '+':
        new_word = message.text[1:].lower()

        wb = load_workbook(filename="words_db.xlsx")
        sheet = wb.active

        eng_words = []
        for row in sheet.rows:
            eng_words.append(str(row[0].value))

        sheet['J1'] = new_word  # Admin ohirgi yozgan so'z
        wb.save('words_db.xlsx')

        if new_word in eng_words:
            await message.answer("Bu so'z mavjud!")
        else:
            await message.answer("Tarjimani yuboring")

    elif message.text[0] == '=':
        trans = message.text[1:]

        wb = load_workbook(filename="words_db.xlsx")
        sheet = wb.active

        new_word = sheet['J1'].value

        eng_words = []

        for row in sheet.rows:
            eng_words.append(str(row[0].value))

        if new_word != None:
            N = sheet.max_row + 1
            sheet[f'A{N}'] = new_word
            sheet[f'B{N}'] = trans
            sheet[f'J1'] = None
            wb.save('words_db.xlsx')
            await message.answer("Yangi so'z qo'shildi!")
        else:
            await message.answer("Avval so'zning inglizchasini +bilan kiriting!")

    elif message.text[0] == '@':
        channel_username = message.text[1:]
        db.delete_channel(channel_username)
        await message.answer(f"{channel_username} kanali majburiy azolikdan chiqarildi")
    else:
        await message.answer('/help_admin')


@dp.message_handler(user_id=ADMINS, content_types=['document'])
async def admin_file(message: types.Message):
    file_name = message.document.file_name
    if file_name == 'words_db.xlsx':
        file_id = message.document.file_id
        file = await dp.bot.get_file(file_id)
        file_path = file.file_path
        await dp.bot.download_file(file_path, file_name)
        await message.answer("So'zlar ombori yangilandi!")
